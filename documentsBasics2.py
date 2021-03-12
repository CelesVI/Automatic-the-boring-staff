import openpyxl,os

wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 42
sheet['A2'] = 'Hello'
os.chdir(r'C:\Users\Bravin\Desktop\Python\Automatic the boring stuff')
wb.save('example.xlsx')
sheet2 = wb.create_sheet()
sheet2.title = 'My new sheet name'
wb.save('example2.xlsx')
wb.create_sheet(index=0, title='My others sheet')
wb.save('example3.xlsx')
